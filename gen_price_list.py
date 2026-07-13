# -*- coding: utf-8 -*-
"""Convert GISTDA Price List PDF (extracted data) to CSV / XLSX / HTML."""
import csv
import html as html_mod
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

OUT_DIR = Path(r"d:\GeoAI")
BASENAME = "Gistda_Price_List"
SOURCE_URL = "https://www.gistda.or.th/download/Gistda_Price_List.pdf"
VAT_NOTE = "**ราคาดังกล่าวยังไม่รวมภาษีมูลค่าเพิ่ม"

NA = "N/A"

# ---------------------------------------------------------------------------
# Data (transcribed from the PDF, 7 pages)
# price keys: archive, tasking, slc, path, monitoring, newacq
# ---------------------------------------------------------------------------

NOTE_A = ("Archive: พื้นที่สั่งขั้นต่ำ 25 ตร.กม. ราคาสำหรับ level Primary (PAN, MS, Pansharpened) | "
          "Tasking: พื้นที่สั่งขั้นต่ำ 100 ตร.กม.")
NOTE_SKYSAT = ("Archive: พื้นที่สั่งขั้นต่ำ 1,250 ตร.กม. เข้าดูผ่าน API/Explorer | "
               "Tasking: โปรดติดต่อเจ้าหน้าที่")
NOTE_B1 = NOTE_A
NOTE_VIDEO = ("ถ่ายภาพเคลื่อนไหวรูปแบบวิดีโอและภาพกลางคืน "
              "ภาพวิดีโอแต่ละช่วงจำกัดความยาว 30 วินาที")
NOTE_NIGHT = ("ถ่ายภาพเคลื่อนไหวรูปแบบวิดีโอและภาพกลางคืน | "
              "พื้นที่สั่งซื้อขั้นต่ำ 100 ตร.กม. ทั้งข้อมูลในคลังและสั่งถ่ายใหม่ | "
              "ความกว้างของแนวถ่ายภาพอย่างน้อย 5 กม.")
NOTE_SPOT = "Archive: พื้นที่สั่งขั้นต่ำ 100 ตร.กม. | Tasking: พื้นที่สั่งขั้นต่ำ 500 ตร.กม."
NOTE_THAICHOTE = "ราคาที่ปรับ Orthorectification แล้ว ราคา 910 บาท/ภาพ"
NOTE_L57 = "คิดเฉพาะค่าดำเนินการผลิตข้อมูลจากคลังข้อมูล | Level 1T"
NOTE_L89 = ("Level 1T มีทั้งหมด 11 Bands | "
            "หากประสงค์ให้ สทอภ. ดำเนินการดาวน์โหลด คิดค่าดำเนินการ 150 บาท/ภาพ")
NOTE_PLANET = ("Access+Download | พื้นที่สั่งขั้นต่ำ 100 ตร.กม. | ระยะเวลาสัญญา 1 ปี | "
               "เข้าดูข้อมูลและดาวน์โหลดผ่าน Planet Explorer, Planet API, Desktop GIS")

SQKM = "บาท/ตร.กม."
IMG = "บาท/ภาพ"

CAT_A = "ดาวเทียมรายละเอียดสูงมาก (30–50 ซม.)"
CAT_B = "ดาวเทียมรายละเอียดสูง (60 ซม. – 2 ม.)"
CAT_C = "ดาวเทียมรายละเอียดปานกลาง (มากกว่า 2 ม.)"
CAT_D = "ดาวเทียมระบบเรดาร์"


def row(cat, sat, mode, res, pol, unit, note, **prices):
    return {
        "cat": cat, "sat": sat, "mode": mode, "res": res, "pol": pol,
        "unit": unit, "note": note,
        "archive": prices.get("archive", ""), "tasking": prices.get("tasking", ""),
        "slc": prices.get("slc", ""), "path": prices.get("path", ""),
        "monitoring": prices.get("monitoring", ""), "newacq": prices.get("newacq", ""),
    }


ROWS = []

# --- Page 1: Very high resolution (30-50 cm), THB/sq.km. ---
for sat, res, ar, ta in [
    ("Pléiades NEO", "30 cm.", 880, 1270),
    ("WorldView-4", "30 cm.", 920, 1560),
    ("SuperView-2", "42 cm.", 700, 1100),
    ("WorldView-1", "50 cm.", 700, 1100),
    ("WorldView-2", "50 cm.", 700, 1100),
    ("WorldView-3", "50 cm.", 700, 1100),
    ("GeoEye-1", "50 cm.", 700, 1100),
    ("Pléiades", "50 cm.", 490, 830),
    ("EarthScanner", "50 cm.", 400, 800),
    ("SuperView-1", "50 cm.", 500, 900),
    ("KOMPSAT-3", "50 cm.", 400, 700),
]:
    ROWS.append(row(CAT_A, sat, "", res, "", SQKM, NOTE_A, archive=ar, tasking=ta))
ROWS.append(row(CAT_A, "SKYSAT", "", "50 cm.", "", SQKM, NOTE_SKYSAT, archive=300, tasking=560))

# --- Page 2-3: High resolution (60 cm - 2 m) ---
for sat, res, ar, ta in [
    ("QuickBird", "60 cm.", 700, NA),
    ("GaoFen-7", "65 cm.", 400, 700),
    ("Jilin", "75 cm.", 300, 600),
    ("DailyVision", "75 cm.", 300, 600),
    ("GaoFen-2", "80 cm.", 300, 400),
    ("IKONOS", "1 m.", 400, NA),
]:
    ROWS.append(row(CAT_B, sat, "", res, "", SQKM, NOTE_B1, archive=ar, tasking=ta))
ROWS.append(row(CAT_B, "Video Constellation", "", "1 m.", "", "บาท/30 วินาที", NOTE_VIDEO,
                archive=142500, tasking=285000))
ROWS.append(row(CAT_B, "Night Imaging", "", "1 m.", "", SQKM, NOTE_NIGHT,
                archive=800, tasking=1400))
ROWS.append(row(CAT_B, "SPOT-6", "", "1.5 m.", "", SQKM, NOTE_SPOT, archive=190, tasking=230))
ROWS.append(row(CAT_B, "SPOT-7", "", "1.5 m.", "", SQKM, NOTE_SPOT, archive=190, tasking=230))
ROWS.append(row(CAT_B, "ไทยโชต", "", "2 m.", "", IMG, NOTE_THAICHOTE, archive=700, tasking=6500))

# --- Page 4: Medium resolution (>2 m) ---
ROWS.append(row(CAT_C, "LANDSAT-5", "", "30 m.", "", IMG, NOTE_L57 + " มีทั้งหมด 7 Bands",
                archive=150, tasking=NA))
ROWS.append(row(CAT_C, "LANDSAT-7", "", "30 m.", "", IMG, NOTE_L57 + " มีทั้งหมด 8 Bands",
                archive=150, tasking=NA))
ROWS.append(row(CAT_C, "LANDSAT-8", "", "30 m.", "", IMG, NOTE_L89, archive=150, tasking=NA))
ROWS.append(row(CAT_C, "LANDSAT-9", "", "30 m.", "", IMG, NOTE_L89, archive=150, tasking=NA))
ROWS.append(row(CAT_C, "PLANETSCOPE", "", "3 m.", "", "บาท/ตร.กม./ปี", NOTE_PLANET,
                archive=180, monitoring=240))

# --- Page 5: RADARSAT-2 (C band), THB/scene, SLC & Path Image ---
for mode, res, slc, path in [
    ("Standard", "25 m.", 57600, 57600),
    ("Spotlight A", "1 m.", 134400, 134400),
    ("Utra-Fine", "3 m.", 86400, 86400),
    ("Wide Utra-Fine", "3 m.", 124800, 124800),
    ("Multi-Look Fine", "8 m.", 67200, 67200),
    ("Wide Multi-Look Fine", "8 m.", 120000, 120000),
    ("Fine", "8 m.", 57600, 57600),
    ("Wide", "30 m.", 57600, 57600),
    ("ScanSAR Narrow", "50 m.", NA, 57600),
    ("ScanSAR Wide", "100 m.", NA, 57600),
    ("Extended High, Low", "25 m.", 57600, 57600),
    ("Fine Quad-Pol", "8 m.", 86400, NA),
    ("Wide Fine Quad-Pol", "8 m.", 124800, NA),
]:
    ROWS.append(row(CAT_D, "RADARSAT-2 (C band)", mode, res, "", IMG, "", slc=slc, path=path))

# --- Page 6: TerraSAR-X (X band) ---
for mode, res, ar, ta in [
    ("Staring Spotlight (ST)", "0.25 m.", 162630, 325260),
    ("High Res Spotlight (HS)", "1 m.", 139230, 278460),
    ("Spotlight", "2 m.", 99450, 198900),
    ("StripMap", "3 m.", 69030, 138060),
    ("ScanSAR", "18.5 m.", 40950, 81900),
    ("Wide ScanSAR", "40 m.", 40950, 81900),
]:
    ROWS.append(row(CAT_D, "TerraSAR-X (X band)", mode, res, "", IMG, "", archive=ar, tasking=ta))

# --- Page 6: COSMO SkyMed (X band), New Acquisition ---
POL_PP = "2 ช่องสัญญาณ polarimattric: HH,VV หรือ HH,HV หรือ VV,VH"
for mode, res, pol, acq in [
    ("Spotlight-2", "1x1 m.", "HH, VV", 180000),
    ("StripMap Himage", "3x3 – 5x5 m.", "HH, HV, VH, VV", 93000),
    ("StripMap PingPong", "10x12 – 20x20 m.", POL_PP, 68000),
    ("ScanSAR Wide", "14x22 – 30x30 m.", "HH, HV, VH, VV", 78000),
    ("ScanSAR Huge", "14x38 – 100x100 m.", "HH, HV, VH, VV", 78000),
]:
    ROWS.append(row(CAT_D, "COSMO SkyMed (X band)", mode, res, pol, IMG, "", newacq=acq))

# --- Page 7: GaoFen-3 (C band) ---
for mode, res, pol, ar, ta in [
    ("Spotlight (SL)", "1 m.", "HH, VV", 116400, 180500),
    ("Ultra-fine Stripmap (UFS)", "3 m.", "HH, VV", 68900, 118800),
    ("Fine Stripmap (FSI)", "5 m.", "HH, VV", 64200, 95000),
    ("Wide Fine Stripmap (FSII)", "10 m.", "HH, HV / VV, VH", 64200, 90300),
    ("Standard Stripmap (SS)", "25 m.", "HH, HV / VV, VH", 54700, 85500),
    ("Narrow ScanSAR (NSC)", "50 m.", "HH, HV / VV, VH", 32100, 42800),
    ("Wide ScanSAR (WSC)", "100 m.", "HH, HV / VV, VH", 32100, 45800),
    ("Quad-pol Stripmap (QPSI)", "8 m.", "HH, HV / VV, VH", 71300, 137800),
    ("Wide Quad-pol Stripmap (QPSII)", "25 m.", "HH, HV / VV, VH", 71300, 137800),
    ("Wave (WAV)", "10 m.", "HH, HV / VV, VH", 10700, 14300),
    ("Global Observation (GLO)", "500 m.", "HH, HV / VV, VH", 10700, 14300),
    ("Extended Incidence Angle (EXT)", "25 m.", "HH, HV / VV, VH", 42800, 57000),
]:
    ROWS.append(row(CAT_D, "GaoFen-3 (C band)", mode, res, pol, IMG, "", archive=ar, tasking=ta))

# ---------------------------------------------------------------------------
# CSV (flat, UTF-8 BOM)
# ---------------------------------------------------------------------------
CSV_HEADERS = [
    "หมวดหมู่", "ดาวเทียม/ระบบ", "โหมด (Mode)", "รายละเอียดภาพ (Resolution)", "Polarization",
    "ราคาในคลัง Standard Archive (บาท)", "ราคาสั่งถ่าย Standard Tasking (บาท)",
    "ราคา Single Look Complex (บาท)", "ราคา Path Image (บาท)",
    "ราคาการติดตาม Monitoring (บาท)", "ราคา New Acquisition (บาท)",
    "หน่วยราคา", "หมายเหตุ",
]
KEYS = ["cat", "sat", "mode", "res", "pol",
        "archive", "tasking", "slc", "path", "monitoring", "newacq", "unit", "note"]


def write_csv():
    with open(OUT_DIR / f"{BASENAME}.csv", "w", newline="", encoding="utf-8-sig") as f:
        w = csv.writer(f)
        w.writerow(CSV_HEADERS)
        for r in ROWS:
            w.writerow([r[k] for k in KEYS])


# ---------------------------------------------------------------------------
# Sub-table specs shared by XLSX and HTML (mirrors the PDF layout per category)
# ---------------------------------------------------------------------------
def sel(cat, sat=None):
    return [r for r in ROWS if r["cat"] == cat and (sat is None or r["sat"] == sat)]


def spec_optical(rows, price1="ข้อมูลในคลัง (Standard Archive)",
                 price2="ข้อมูลชนิดสั่งถ่าย (Standard Tasking)", k1="archive", k2="tasking"):
    return {
        "headers": ["ดาวเทียม (Satellite)", "รายละเอียดภาพ (Resolution)", price1, price2],
        "rows": [[r["sat"] or r["mode"], r["res"], r[k1], r[k2]] for r in rows],
    }


CATEGORIES = [
    {
        "title": f"ราคาข้อมูลจากดาวเทียมรายละเอียดสูงมาก (รายละเอียดภาพ 30–50 ซม.)",
        "sheet": "สูงมาก 30-50 ซม.",
        "subtables": [
            dict(spec_optical([r for r in sel(CAT_A) if r["sat"] != "SKYSAT"]),
                 unit=f"หน่วย : {SQKM}", notes=[NOTE_A]),
            dict(spec_optical(sel(CAT_A, "SKYSAT")), unit=f"หน่วย : {SQKM}", notes=[NOTE_SKYSAT]),
        ],
    },
    {
        "title": "ราคาข้อมูลจากดาวเทียมรายละเอียดสูง (รายละเอียดภาพ 60 ซม. – 2 ม.)",
        "sheet": "สูง 60 ซม.-2 ม.",
        "subtables": [
            dict(spec_optical([r for r in sel(CAT_B)
                               if r["sat"] in ("QuickBird", "GaoFen-7", "Jilin",
                                               "DailyVision", "GaoFen-2", "IKONOS")]),
                 unit=f"หน่วย : {SQKM}", notes=[NOTE_B1]),
            dict(spec_optical(sel(CAT_B, "Video Constellation")),
                 unit="หน่วย : บาท/30 วินาที", notes=[NOTE_VIDEO]),
            dict(spec_optical(sel(CAT_B, "Night Imaging")),
                 unit=f"หน่วย : {SQKM}", notes=[NOTE_NIGHT]),
            dict(spec_optical([r for r in sel(CAT_B) if r["sat"].startswith("SPOT")]),
                 unit=f"หน่วย : {SQKM}", notes=[NOTE_SPOT]),
            dict(spec_optical(sel(CAT_B, "ไทยโชต")),
                 unit=f"หน่วย : {IMG}", notes=[NOTE_THAICHOTE]),
        ],
    },
    {
        "title": "ราคาข้อมูลจากดาวเทียมรายละเอียดปานกลาง (รายละเอียดมากกว่า 2 เมตร)",
        "sheet": "ปานกลาง มากกว่า 2 ม.",
        "subtables": [
            dict(spec_optical([r for r in sel(CAT_C) if r["sat"].startswith("LANDSAT")]),
                 unit=f"หน่วย : {IMG}",
                 notes=["คิดเฉพาะค่าดำเนินการผลิตข้อมูลจากคลังข้อมูล",
                        "LANDSAT-5: Level 1T มีทั้งหมด 7 Bands / LANDSAT-7: Level 1T มีทั้งหมด 8 Bands",
                        "LANDSAT-8, LANDSAT-9: Level 1T มีทั้งหมด 11 Bands "
                        "หากประสงค์ให้ สทอภ. ดำเนินการดาวน์โหลด คิดค่าดำเนินการ 150 บาท/ภาพ"]),
            {
                "headers": ["ดาวเทียม (Satellite)", "รายละเอียดภาพ (Resolution)", "รูปแบบ",
                            "ข้อมูลในคลัง (Standard Archive)", "การติดตาม (Monitoring)"],
                "rows": [["PLANETSCOPE", "3 m.", "Access+Download", 180, 240]],
                "unit": "หน่วย : บาท/ตร.กม./ปี",
                "notes": [NOTE_PLANET],
            },
        ],
    },
    {
        "title": "ราคาข้อมูลจากดาวเทียมระบบเรดาร์",
        "sheet": "เรดาร์",
        "subtables": [
            {
                "caption": "RADARSAT-2 (C band)",
                "headers": ["Mode", "รายละเอียดภาพ (Resolution)",
                            "Single Look Complex (บาท)", "Path Image (บาท)"],
                "rows": [[r["mode"], r["res"], r["slc"], r["path"]]
                         for r in sel(CAT_D, "RADARSAT-2 (C band)")],
                "unit": f"หน่วย : {IMG}", "notes": [],
            },
            {
                "caption": "TerraSAR-X (X band)",
                "headers": ["Mode", "รายละเอียดภาพ (Resolution)",
                            "ข้อมูลในคลัง (Standard Archive)", "ข้อมูลชนิดสั่งถ่าย (Standard Tasking)"],
                "rows": [[r["mode"], r["res"], r["archive"], r["tasking"]]
                         for r in sel(CAT_D, "TerraSAR-X (X band)")],
                "unit": f"หน่วย : {IMG}", "notes": [],
            },
            {
                "caption": "COSMO SkyMed (X band)",
                "headers": ["Mode", "รายละเอียดภาพ (Resolution)", "Polarization", "New Acquisition (บาท)"],
                "rows": [[r["mode"], r["res"], r["pol"], r["newacq"]]
                         for r in sel(CAT_D, "COSMO SkyMed (X band)")],
                "unit": f"หน่วย : {IMG}", "notes": [],
            },
            {
                "caption": "GaoFen-3 (C band)",
                "headers": ["Mode", "รายละเอียดภาพ (Resolution)", "Polarization",
                            "ข้อมูลในคลัง (Standard Archive)", "ข้อมูลชนิดสั่งถ่าย (Standard Tasking)"],
                "rows": [[r["mode"], r["res"], r["pol"], r["archive"], r["tasking"]]
                         for r in sel(CAT_D, "GaoFen-3 (C band)")],
                "unit": f"หน่วย : {IMG}", "notes": [],
            },
        ],
    },
]

# ---------------------------------------------------------------------------
# XLSX
# ---------------------------------------------------------------------------
def write_xlsx():
    wb = Workbook()
    wb.remove(wb.active)

    head_fill = PatternFill("solid", fgColor="CDEBF5")
    cap_fill = PatternFill("solid", fgColor="D9D9D9")
    thin = Side(style="thin", color="808080")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    bold = Font(bold=True)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left = Alignment(horizontal="left", vertical="center", wrap_text=True)

    for cat in CATEGORIES:
        ws = wb.create_sheet(cat["sheet"])
        ncols = max(len(t["headers"]) for t in cat["subtables"])
        r_i = 1
        ws.merge_cells(start_row=r_i, start_column=1, end_row=r_i, end_column=ncols)
        c = ws.cell(r_i, 1, cat["title"])
        c.font = Font(bold=True, size=14)
        r_i += 2

        for t in cat["subtables"]:
            w = len(t["headers"])
            ws.merge_cells(start_row=r_i, start_column=1, end_row=r_i, end_column=w)
            ws.cell(r_i, 1, t.get("unit", "")).alignment = Alignment(horizontal="right")
            r_i += 1
            if t.get("caption"):
                ws.merge_cells(start_row=r_i, start_column=1, end_row=r_i, end_column=w)
                cc = ws.cell(r_i, 1, t["caption"])
                cc.font = bold
                cc.fill = cap_fill
                cc.alignment = center
                cc.border = border
                r_i += 1
            for j, h in enumerate(t["headers"], 1):
                hc = ws.cell(r_i, j, h)
                hc.font = bold
                hc.fill = head_fill
                hc.alignment = center
                hc.border = border
            r_i += 1
            for data_row in t["rows"]:
                for j, v in enumerate(data_row, 1):
                    dc = ws.cell(r_i, j, v)
                    dc.border = border
                    if isinstance(v, (int, float)):
                        dc.number_format = "#,##0"
                        dc.alignment = Alignment(horizontal="right", vertical="center")
                    else:
                        dc.alignment = left if j == 1 else center
                r_i += 1
            for note in t["notes"]:
                ws.merge_cells(start_row=r_i, start_column=1, end_row=r_i, end_column=w)
                nc = ws.cell(r_i, 1, "• " + note)
                nc.font = Font(size=9, italic=True)
                nc.alignment = left
                r_i += 1
            r_i += 1

        ws.cell(r_i, 1, VAT_NOTE).font = Font(bold=True, color="C00000")
        widths = [34, 22, 26, 26, 26]
        for j in range(1, ncols + 1):
            ws.column_dimensions[get_column_letter(j)].width = widths[min(j - 1, len(widths) - 1)]

    # Flat sheet mirroring the CSV
    ws = wb.create_sheet("รวม (Flat)")
    for j, h in enumerate(CSV_HEADERS, 1):
        hc = ws.cell(1, j, h)
        hc.font = bold
        hc.fill = head_fill
        hc.alignment = center
        hc.border = border
    for i, r in enumerate(ROWS, 2):
        for j, k in enumerate(KEYS, 1):
            v = r[k]
            dc = ws.cell(i, j, v if v != "" else None)
            dc.border = border
            if isinstance(v, (int, float)):
                dc.number_format = "#,##0"
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(CSV_HEADERS))}{len(ROWS) + 1}"
    for j, wdt in enumerate([38, 24, 28, 20, 26, 16, 16, 14, 14, 14, 14, 16, 60], 1):
        ws.column_dimensions[get_column_letter(j)].width = wdt

    wb.save(OUT_DIR / f"{BASENAME}.xlsx")


# ---------------------------------------------------------------------------
# HTML
# ---------------------------------------------------------------------------
def esc(v):
    return html_mod.escape(str(v))


def fmt(v):
    return f"{v:,}" if isinstance(v, (int, float)) else esc(v)


def write_html():
    parts = ["""<!DOCTYPE html>
<html lang="th">
<head>
<meta charset="utf-8">
<title>GISTDA Price List — ราคาข้อมูลจากดาวเทียม</title>
<style>
  body { font-family: "Segoe UI", Tahoma, sans-serif; margin: 2rem auto; max-width: 960px;
         color: #1a2b3c; background: #fff; padding: 0 1rem; }
  h1 { color: #0b5e8a; border-bottom: 3px solid #0b5e8a; padding-bottom: .4rem; }
  h2 { color: #0b5e8a; margin-top: 2.2rem; }
  table { border-collapse: collapse; width: 100%; margin: .6rem 0 .3rem; font-size: .95rem; }
  caption { caption-side: top; text-align: left; font-weight: bold; background: #d9d9d9;
            padding: .4rem .6rem; border: 1px solid #999; border-bottom: none; }
  th { background: #cdebf5; padding: .5rem .6rem; border: 1px solid #999; }
  td { padding: .4rem .6rem; border: 1px solid #999; }
  td.num { text-align: right; white-space: nowrap; }
  tr:nth-child(even) td { background: #f4fafd; }
  .unit { text-align: right; font-size: .85rem; color: #555; margin: .8rem 0 0; }
  ul.notes { font-size: .85rem; color: #444; margin: .2rem 0 1rem; }
  .vat { color: #c00000; font-weight: bold; margin-top: 1.5rem; }
  footer { margin-top: 2.5rem; padding-top: 1rem; border-top: 1px solid #ccc;
           font-size: .85rem; color: #555; }
</style>
</head>
<body>
<h1>GISTDA Price List — ราคาข้อมูลจากดาวเทียม</h1>
"""]
    for cat in CATEGORIES:
        parts.append(f"<h2>{esc(cat['title'])}</h2>\n")
        for t in cat["subtables"]:
            parts.append(f'<p class="unit">{esc(t.get("unit", ""))}</p>\n<table>\n')
            if t.get("caption"):
                parts.append(f"<caption>{esc(t['caption'])}</caption>\n")
            parts.append("<thead><tr>" +
                         "".join(f"<th>{esc(h)}</th>" for h in t["headers"]) +
                         "</tr></thead>\n<tbody>\n")
            for data_row in t["rows"]:
                cells = []
                for v in data_row:
                    cls = ' class="num"' if isinstance(v, (int, float)) else ""
                    cells.append(f"<td{cls}>{fmt(v)}</td>")
                parts.append("<tr>" + "".join(cells) + "</tr>\n")
            parts.append("</tbody></table>\n")
            if t["notes"]:
                parts.append('<ul class="notes">' +
                             "".join(f"<li>{esc(n)}</li>" for n in t["notes"]) + "</ul>\n")
    parts.append(f'<p class="vat">{esc(VAT_NOTE)}</p>\n')
    parts.append(f"""<footer>
<p><strong>สอบถามรายละเอียดเพิ่มเติมได้ที่</strong><br>
ฝ่ายพัฒนาธุรกิจและการบริการ สำนักงานพัฒนาเทคโนโลยีอวกาศและภูมิสารสนเทศ (องค์การมหาชน)<br>
ศูนย์ราชการเฉลิมพระเกียรติ อาคาร B ชั้น 6 ถนนแจ้งวัฒนะ แขวงทุ่งสองห้อง เขตหลักสี่ กรุงเทพฯ 10210<br>
โทร 0 2141 4564-66, 69 &nbsp;|&nbsp; โทรสาร 0 2143 9593 &nbsp;|&nbsp; อีเมล usd@gistda.or.th</p>
<p>ที่มา: <a href="{SOURCE_URL}">{SOURCE_URL}</a></p>
</footer>
</body>
</html>
""")
    (OUT_DIR / f"{BASENAME}.html").write_text("".join(parts), encoding="utf-8")


if __name__ == "__main__":
    write_csv()
    write_xlsx()
    write_html()
    print(f"rows total: {len(ROWS)}")
    for c in (CAT_A, CAT_B, CAT_C, CAT_D):
        print(f"  {c}: {len(sel(c))}")
    for ext in ("csv", "xlsx", "html"):
        p = OUT_DIR / f"{BASENAME}.{ext}"
        print(p, p.stat().st_size, "bytes")
